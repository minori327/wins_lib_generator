#!/usr/bin/env python3
"""
Story Identifier v2.0 - Work Stream 2

Generate candidate Success Story rows for human review.

HUMAN PRIMACY ENFORCEMENT:
- NO consolidation
- NO deduplication
- NO final business judgment
- Generate candidates only
- Provide consolidation hints (textual, advisory)
"""

import logging
import json
import yaml
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime


logger = logging.getLogger(__name__)


def load_prompt_template(template_path: str) -> str:
    """Load prompt template from YAML file.

    Args:
        template_path: Path to prompt template YAML file

    Returns:
        Prompt template string

    Raises:
        FileNotFoundError: If template file doesn't exist
    """
    template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Prompt template not found: {template_path}")

    with open(template_path, 'r') as f:
        data = yaml.safe_load(f)

    return data.get('prompt', '')


def discover_markdown_files(sources_path: Path) -> List[Path]:
    """Discover all Markdown files in sources directory.

    Args:
        sources_path: Path to Markdown sources directory

    Returns:
        List of Markdown file paths
    """
    logger.info(f"[DISCOVER_MD] Scanning: {sources_path}")

    files = list(sources_path.glob("*.md"))

    logger.info(f"[DISCOVER_MD] Found {len(files)} Markdown files")
    return files


def parse_source_id_from_markdown(file_path: Path) -> Optional[str]:
    """Extract source_id from Markdown frontmatter.

    Args:
        file_path: Path to Markdown file

    Returns:
        source_id string or None
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse frontmatter
    if content.startswith('---'):
        # Find end of frontmatter
        end_idx = content.find('---', 3)
        if end_idx != -1:
            frontmatter_text = content[3:end_idx]
            try:
                frontmatter = yaml.safe_load(frontmatter_text)
                return frontmatter.get('source_id')
            except yaml.YAMLError:
                logger.warning(f"[PARSE_MD] Failed to parse frontmatter in {file_path}")

    return None


def extract_candidates_from_source(
    markdown_path: Path,
    source_id: str,
    prompt_template: str,
    llm_client: Any
) -> List[Dict[str, Any]]:
    """Extract candidate Success Stories from a single Markdown source.

    Args:
        markdown_path: Path to Markdown file
        source_id: Source ID
        prompt_template: Prompt template string
        llm_client: LLM client instance

    Returns:
        List of candidate story dictionaries

    Raises:
        RuntimeError: If LLM call fails
    """
    logger.info(f"[EXTRACT_CANDIDATES] Processing source: {source_id}")

    # Read Markdown content
    with open(markdown_path, 'r', encoding='utf-8') as f:
        source_text = f.read()

    # Truncate if too long (for LLM context window)
    max_length = 10000  # Conservative limit
    if len(source_text) > max_length:
        logger.warning(f"[EXTRACT_CANDIDATES] Truncating source from {len(source_text)} to {max_length} characters")
        source_text = source_text[:max_length] + "\n\n[Content truncated due to length...]"

    # Build prompt
    prompt = prompt_template.replace("{{source_text}}", source_text)
    prompt = prompt.replace("{{source_id}}", source_id)

    # Call LLM
    logger.info(f"[EXTRACT_CANDIDATES] Calling LLM for source: {source_id}")
    try:
        response = llm_client.call_json(prompt)
    except Exception as e:
        error_msg = f"LLM call failed for source {source_id}: {e}"
        logger.error(f"[EXTRACT_CANDIDATES] {error_msg}")
        raise RuntimeError(error_msg) from e

    # Parse response
    if isinstance(response, list):
        candidates = response
    elif isinstance(response, dict):
        # LLM returned a dict with a 'stories' key or similar
        if 'candidates' in response:
            candidates = response['candidates']
        elif 'stories' in response:
            candidates = response['stories']
        else:
            candidates = [response]
    else:
        error_msg = f"Unexpected LLM response type: {type(response)}"
        logger.error(f"[EXTRACT_CANDIDATES] {error_msg}")
        raise RuntimeError(error_msg)

    logger.info(f"[EXTRACT_CANDIDATES] Extracted {len(candidates)} candidates from {source_id}")

    # Add metadata to each candidate
    for candidate in candidates:
        candidate['source_id'] = source_id
        candidate['source_file'] = str(markdown_path)

    return candidates


def normalize_candidate(candidate: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize candidate data structure.

    Args:
        candidate: Raw candidate dictionary from LLM

    Returns:
        Normalized candidate dictionary
    """
    return {
        'status': 'candidate',  # All generated rows are candidates
        'customer': candidate.get('customer', ''),
        'context': candidate.get('context', ''),
        'action': candidate.get('action', ''),
        'outcome': candidate.get('outcome', ''),
        'metrics': json.dumps(candidate.get('metrics', [])),  # Store as JSON string
        'evidence_references': json.dumps(candidate.get('evidence_references', [])),
        'consolidation_hints': candidate.get('consolidation_hints', ''),
        'source_id': candidate.get('source_id', ''),
        'source_file': candidate.get('source_file', ''),
    }


def create_excel_workbook(candidates: List[Dict[str, Any]], output_path: Path) -> None:
    """Create Excel workbook with candidate stories.

    Work Stream 3: Excel as consolidation workspace.

    Args:
        candidates: List of normalized candidate dictionaries
        output_path: Path to output Excel file

    Raises:
        RuntimeError: If Excel creation fails
    """
    logger.info(f"[CREATE_EXCEL] Creating workbook: {output_path}")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel output. Install: pip install openpyxl")

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create worksheets
    ws_stories = wb.create_sheet("Stories", 0)
    ws_evidence = wb.create_sheet("Evidence", 1)
    ws_metrics = wb.create_sheet("Metrics", 2)
    ws_metadata = wb.create_sheet("Metadata", 3)

    # Define column widths
    ws_stories.column_dimensions['A'].width = 15  # status
    ws_stories.column_dimensions['B'].width = 25  # customer
    ws_stories.column_dimensions['C'].width = 30  # context
    ws_stories.column_dimensions['D'].width = 30  # action
    ws_stories.column_dimensions['E'].width = 30  # outcome
    ws_stories.column_dimensions['F'].width = 40  # metrics
    ws_stories.column_dimensions['G'].width = 50  # evidence_references
    ws_stories.column_dimensions['H'].width = 50  # consolidation_hints
    ws_stories.column_dimensions['I'].width = 20  # source_id

    # Stories worksheet header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    headers = [
        'status', 'customer', 'context', 'action', 'outcome',
        'metrics', 'evidence_references', 'consolidation_hints', 'source_id'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_stories.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Populate Stories worksheet
    for row_idx, candidate in enumerate(candidates, 2):
        ws_stories.cell(row=row_idx, column=1, value=candidate['status'])
        ws_stories.cell(row=row_idx, column=2, value=candidate['customer'])
        ws_stories.cell(row=row_idx, column=3, value=candidate['context'])
        ws_stories.cell(row=row_idx, column=4, value=candidate['action'])
        ws_stories.cell(row=row_idx, column=5, value=candidate['outcome'])
        ws_stories.cell(row=row_idx, column=6, value=candidate['metrics'])
        ws_stories.cell(row=row_idx, column=7, value=candidate['evidence_references'])
        ws_stories.cell(row=row_idx, column=8, value=candidate['consolidation_hints'])
        ws_stories.cell(row=row_idx, column=9, value=candidate['source_id'])

    # Evidence worksheet header
    ws_evidence.cell(row=1, column=1, value='story_id').font = header_font
    ws_evidence.cell(row=1, column=2, value='source_id').font = header_font
    ws_evidence.cell(row=1, column=3, value='eference').font = header_font

    # Populate Evidence worksheet (simplified - each row gets its evidence)
    for row_idx, candidate in enumerate(candidates, 2):
        ws_evidence.cell(row=row_idx, column=1, value=f"row-{row_idx}")
        ws_evidence.cell(row=row_idx, column=2, value=candidate['source_id'])
        ws_evidence.cell(row=row_idx, column=3, value=candidate['evidence_references'])

    # Metrics worksheet header
    ws_metrics.cell(row=1, column=1, value='story_id').font = header_font
    ws_metrics.cell(row=1, column=2, value='metric').font = header_font

    # Populate Metrics worksheet (parse JSON metrics)
    metric_row = 2
    for row_idx, candidate in enumerate(candidates, 2):
        story_id = f"row-{row_idx}"
        try:
            metrics_list = json.loads(candidate['metrics'])
            for metric in metrics_list:
                ws_metrics.cell(row=metric_row, column=1, value=story_id)
                ws_metrics.cell(row=metric_row, column=2, value=str(metric))
                metric_row += 1
        except json.JSONDecodeError:
            # If metrics are not valid JSON, store as-is
            ws_metrics.cell(row=metric_row, column=1, value=story_id)
            ws_metrics.cell(row=metric_row, column=2, value=candidate['metrics'])
            metric_row += 1

    # Metadata worksheet
    ws_metadata.column_dimensions['A'].width = 20
    ws_metadata.column_dimensions['B'].width = 50

    ws_metadata.cell(row=1, column=1, value="Key").font = header_font
    ws_metadata.cell(row=1, column=2, value="Value").font = header_font

    # Generate version number (find existing files)
    version = 1
    if output_path.parent.exists():
        existing = list(output_path.parent.glob("candidates_v*_*.xlsx"))
        if existing:
            # Extract highest version number
            versions = []
            for f in existing:
                try:
                    # Extract version from filename: candidates_vN_YYYYMMDD.xlsx
                    name = f.stem
                    parts = name.split('_')
                    if len(parts) >= 2 and parts[0].startswith('candidates_v'):
                        v = int(parts[0][11:])  # Remove 'candidates_v' prefix
                        versions.append(v)
                except (ValueError, IndexError):
                    pass
            if versions:
                version = max(versions) + 1

    # Generate datestamp
    datestamp = datetime.utcnow().strftime('%Y%m%d')

    # Set metadata
    metadata_rows = [
        ("version", version),
        ("generated_at", datetime.utcnow().isoformat() + "Z"),
        ("generated_by", "system"),
        ("total_candidates", len(candidates)),
        ("datestamp", datestamp),
    ]

    for row_idx, (key, value) in enumerate(metadata_rows, 2):
        ws_metadata.cell(row=row_idx, column=1, value=key)
        ws_metadata.cell(row=row_idx, column=2, value=str(value))

    # Save workbook
    # Ensure filename includes version
    filename = f"candidates_v{version}_{datestamp}.xlsx"
    final_output_path = output_path.parent / filename

    wb.save(final_output_path)

    logger.info(f"[CREATE_EXCEL] Saved workbook: {final_output_path}")
    logger.info(f"[CREATE_EXCEL] Version: {version}, Candidates: {len(candidates)}")


def identify_stories(
    sources_path: Path,
    output_path: Path,
    config: Dict[str, Any]
) -> None:
    """Identify Success Story candidates from Markdown sources.

    Main entry point for Work Stream 2.

    Args:
        sources_path: Path to Markdown sources directory
        output_path: Path to output Excel file
        config: Configuration dictionary

    Raises:
        RuntimeError: If critical failures occur

    HUMAN PRIMACY ENFORCEMENT:
    - This function generates candidates ONLY
    - NO consolidation logic
    - NO deduplication logic
    - NO automatic status changes
    """
    logger.info(f"[IDENTIFY_STORIES] Starting operation...")
    logger.info(f"[IDENTIFY_STORIES] Sources: {sources_path}")
    logger.info(f"[IDENTIFY_STORIES] Output: {output_path}")

    # Load LLM client
    from utils.llm_client import LLMClient
    llm_client = LLMClient(config)

    # Load prompt template
    prompt_template_path = Path(config["paths"]["prompts_dir"]) / "identify_stories.yaml"
    prompt_template = load_prompt_template(prompt_template_path)

    # Discover Markdown files
    md_files = discover_markdown_files(sources_path)

    if not md_files:
        logger.warning(f"[IDENTIFY_STORIES] No Markdown files found in {sources_path}")
        return

    # Extract candidates from each source
    all_candidates = []
    success_count = 0
    fail_count = 0

    for md_file in md_files:
        try:
            # Parse source_id
            source_id = parse_source_id_from_markdown(md_file)
            if not source_id:
                logger.warning(f"[IDENTIFY_STORIES] No source_id found in {md_file}, skipping")
                continue

            # Extract candidates
            candidates = extract_candidates_from_source(
                markdown_path=md_file,
                source_id=source_id,
                prompt_template=prompt_template,
                llm_client=llm_client
            )

            # Normalize and add to list
            for candidate in candidates:
                normalized = normalize_candidate(candidate)
                all_candidates.append(normalized)

            success_count += 1

        except Exception as e:
            logger.error(f"[IDENTIFY_STORIES] Failed to process {md_file}: {e}")
            fail_count += 1
            continue

    # Summary
    logger.info(f"[IDENTIFY_STORIES] Extraction complete: {success_count} succeeded, {fail_count} failed")
    logger.info(f"[IDENTIFY_STORIES] Total candidates generated: {len(all_candidates)}")

    if not all_candidates:
        logger.warning(f"[IDENTIFY_STORIES] No candidates generated")
        return

    # Create Excel workbook
    create_excel_workbook(all_candidates, output_path)

    logger.info(f"[IDENTIFY_STORIES] Completed successfully")
