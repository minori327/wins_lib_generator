#!/usr/bin/env python3
"""
Summary Generator v2.0 - Work Stream 4

Generate executive summaries from final Excel rows.

HUMAN PRIMACY ENFORCEMENT:
- ONLY final rows are consumed
- NO ambiguity resolution
- NO content rewriting
- Reflects human-approved content only
"""

import logging
import json
import yaml
from pathlib import Path
from typing import Dict, Any, List


logger = logging.getLogger(__name__)


def load_prompt_template(template_path: str) -> str:
    """Load prompt template from YAML file.

    Args:
        template_path: Path to prompt template YAML file

    Returns:
        Prompt template string

    Raises:
        FileNotFoundError: If template file doesn't exist
    """
    template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Prompt template not found: {template_path}")

    with open(template_path, 'r') as f:
        data = yaml.safe_load(f)

    return data.get('prompt', '')


def load_final_stories_from_excel(input_path: Path) -> List[Dict[str, Any]]:
    """Load ONLY final stories from Excel workbook.

    Args:
        input_path: Path to Excel workbook

    Returns:
        List of final story dictionaries

    Raises:
        RuntimeError: If Excel loading fails
    """
    logger.info(f"[LOAD_EXCEL] Loading final stories from: {input_path}")

    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel input. Install: pip install openpyxl")

    if not input_path.exists():
        raise FileNotFoundError(f"Excel file not found: {input_path}")

    wb = openpyxl.load_workbook(input_path)

    if "Stories" not in wb.sheetnames:
        raise RuntimeError(f"Excel file missing 'Stories' worksheet: {input_path}")

    ws = wb["Stories"]

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    logger.debug(f"[LOAD_EXCEL] Headers: {headers}")

    # Find column indices
    try:
        status_col = headers.index('status') + 1
        customer_col = headers.index('customer') + 1
        context_col = headers.index('context') + 1
        action_col = headers.index('action') + 1
        outcome_col = headers.index('outcome') + 1
        metrics_col = headers.index('metrics') + 1
        evidence_col = headers.index('evidence_references') + 1
    except ValueError as e:
        raise RuntimeError(f"Excel file missing required column: {e}")

    # Load final rows only
    final_stories = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        status = row[status_col - 1].value

        # Skip non-final rows
        if status != 'final':
            logger.debug(f"[LOAD_EXCEL] Skipping row {row_idx}: status={status}")
            continue

        # Extract story data
        story = {
            'customer': row[customer_col - 1].value or '',
            'context': row[context_col - 1].value or '',
            'action': row[action_col - 1].value or '',
            'outcome': row[outcome_col - 1].value or '',
            'metrics': row[metrics_col - 1].value or '',
            'evidence_references': row[evidence_col - 1].value or '',
        }

        final_stories.append(story)
        logger.debug(f"[LOAD_EXCEL] Loaded final story from row {row_idx}: {story['customer']}")

    logger.info(f"[LOAD_EXCEL] Loaded {len(final_stories)} final stories")
    return final_stories


def format_stories_for_prompt(stories: List[Dict[str, Any]]) -> str:
    """Format stories for LLM prompt.

    Args:
        stories: List of story dictionaries

    Returns:
        Formatted string for prompt
    """
    formatted_stories = []

    for idx, story in enumerate(stories, 1):
        # Parse metrics JSON if possible
        try:
            metrics = json.loads(story['metrics'])
            metrics_str = ', '.join(str(m) for m in metrics)
        except (json.JSONDecodeError, TypeError):
            metrics_str = str(story['metrics'])

        story_text = f"""
Story {idx}:
  Customer: {story['customer']}
  Context: {story['context']}
  Action: {story['action']}
  Outcome: {story['outcome']}
  Metrics: {metrics_str}
  Evidence: {story['evidence_references']}
"""
        formatted_stories.append(story_text)

    return "\n".join(formatted_stories)


def generate_summary_with_llm(
    stories_text: str,
    prompt_template: str,
    llm_client: Any
) -> str:
    """Generate summary using LLM.

    Args:
        stories_text: Formatted stories text
        prompt_template: Prompt template
        llm_client: LLM client instance

    Returns:
        Generated summary text

    Raises:
        RuntimeError: If LLM call fails
    """
    logger.info(f"[GEN_SUMMARY] Calling LLM to generate summary...")

    # Build prompt
    prompt = prompt_template.replace("{{stories}}", stories_text)

    # Call LLM
    try:
        response = llm_client.call(prompt, output_format="markdown")
    except Exception as e:
        error_msg = f"LLM call failed: {e}"
        logger.error(f"[GEN_SUMMARY] {error_msg}")
        raise RuntimeError(error_msg) from e

    logger.info(f"[GEN_SUMMARY] Generated summary: {len(response)} characters")
    return response


def add_traceability_footer(summary: str, input_path: Path) -> str:
    """Add traceability footer to summary.

    Args:
        summary: Summary text
        input_path: Input Excel file path

    Returns:
        Summary with footer
    """
    footer = f"""

---

**Traceability**
- Source: {input_path.name}
- Generated: {datetime.now().isoformat()}Z
- Stories: All have status=final (human-approved)

This summary reflects ONLY human-approved content from the Excel workbook.
"""
    from datetime import datetime

    return summary + footer


def generate_summary(
    input_path: Path,
    output_path: Path,
    prompt_template: str,
    config: Dict[str, Any]
) -> None:
    """Generate executive summary from final Excel rows.

    Main entry point for Work Stream 4.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output Markdown file
        prompt_template: Path to prompt template
        config: Configuration dictionary

    Raises:
        RuntimeError: If critical failures occur

    HUMAN PRIMACY ENFORCEMENT:
    - ONLY final rows are consumed
    - NO ambiguity resolution
    - NO content rewriting
    """
    logger.info(f"[GENERATE_SUMMARY] Starting operation...")
    logger.info(f"[GENERATE_SUMMARY] Input: {input_path}")
    logger.info(f"[GENERATE_SUMMARY] Output: {output_path}")
    logger.info(f"[GENERATE_SUMMARY] Prompt: {prompt_template}")

    # Load LLM client
    from utils.llm_client import LLMClient
    llm_client = LLMClient(config)

    # Load prompt template
    prompt_text = load_prompt_template(prompt_template)

    # Load final stories from Excel
    final_stories = load_final_stories_from_excel(input_path)

    if not final_stories:
        warning_msg = (
            f"[GENERATE_SUMMARY] WARNING: No final stories found in {input_path}\n"
            f"[GENERATE_SUMMARY] Only rows with status='final' are consumed.\n"
            f"[GENERATE_SUMMARY] Please review the Excel file and set appropriate rows to 'final'."
        )
        logger.warning(warning_msg)

        # Generate warning summary
        summary = f"""# Executive Summary

**WARNING**: No stories with status='final' were found in the input file.

To generate a summary, please:
1. Open the Excel file: {input_path}
2. Review the candidate stories
3. Set approved stories to status='final'
4. Re-run this command

This system ONLY consumes human-approved content (status='final').
"""
        # Write output
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(summary)

        logger.info(f"[GENERATE_SUMMARY] Generated warning summary: {output_path}")
        return

    # Format stories for prompt
    stories_text = format_stories_for_prompt(final_stories)

    # Generate summary
    summary = generate_summary_with_llm(
        stories_text=stories_text,
        prompt_template=prompt_text,
        llm_client=llm_client
    )

    # Add traceability footer
    summary = add_traceability_footer(summary, input_path)

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(summary)

    logger.info(f"[GENERATE_SUMMARY] Completed successfully: {output_path}")
